#This code is taking from a specific range of columns their values, delete all n/a and null values
#and then add it to a user specify columns in excel
#But as con, the code doesnt save the formulaes in a new file
#so this should be used as a endway editing

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

#path = 'C:\\Users\\ASpiridonov\\Desktop\\study\\loader.xlsm'

wb = load_workbook('loader.xlsm', data_only = True)

sheet_obj = wb.active

max_row = sheet_obj.max_column

valuelist = list()

#this is is getting values from a column range and 
#appends not nulls or n/a to a list
for i in range (1, 707):
	cell_obj = sheet_obj.cell ( column = 23, row = i)
	valuelist.append(cell_obj.value)
values = list()
for value in valuelist:
	if type(value) is int:
		values.append(value)
	else:
		continue

#print(values) #this one checks if i get the list right
k=15
y=24
#this one is takin in consideration the k and y values and 
#insert values in a previous list into the specific column one after one
for i, value in enumerate(values):
	sheet_obj.cell(column=y, row=k+i, value=value)
wb.save('loadermodified.xlsx')
print('Выполнено, вы можете найти результат в файле с названием loadermodified.xlsx')